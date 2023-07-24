import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Get the number of rows in the spreadsheet
# Start on the 2nd row and go through max + 1 because range doesn't include the last calcualted row
for product_row in range(2, product_list.max_row + 1):
    # Get the supplier name
    # Each cell in column 4
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)


    # Calculate number of products per supplier
    # Check to see if supplier is already in the dictionary(key,value) named products_per_supplier
    # We're also incrementing the number of products each supplier has
    if supplier_name in products_per_supplier:
        # How many products does the supplier have?
        # Increment current_num_products each time the supplier name shows up in the next row
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    # For the first time a new supplier is seen in the list, set the products to 1
    # This would happen on any new supplier and the first time the program is run since products_per_supplier is empty
    else:
        products_per_supplier[supplier_name] = 1


    # Calculate total value of inventory per supplier
    # Same logic as above
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price


    # Logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)


    # Add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")
